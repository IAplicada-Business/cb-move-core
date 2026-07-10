#!/usr/bin/env python3
"""Aplica dados de docs/docs2: IM Focus, convênios e e-mails de pacientes."""
from __future__ import annotations

import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.request
from pathlib import Path

from load_app_env import load_app_env

IM_CB_MOVE = "1477199"
XLSX_PATH = Path(__file__).resolve().parent.parent / "docs" / "docs2" / "Lista_pacientes_email.xlsx"

CONVENIOS = [
    {
        "match": "bradesco",
        "nome": "Bradesco Seguros",
        "cnpj": "92693118000160",
        "razao_social": "BRADESCO SAUDE S/A",
        "email_nf": "liminarprestador@bradescoseguros.com.br",
        "endereco": "AV RIO DE JANEIRO, 555, SAL 801-SAL 1701, CAJU",
        "cep": "20931675",
        "codigo_municipio_ibge": 3304557,
        "cidade": "Rio de Janeiro",
        "uf": "RJ",
    },
    {
        "match": "centro",
        "nome": "Centro Clínico Gaúcho",
        "cnpj": "00773639000100",
        "razao_social": "CENTRO CLINICO GAUCHO LTDA",
        "email_nf": "extra.folha@ccgrs.com.br",
        "endereco": "AV HERACLITO GRACA, 406, CENTRO",
        "cep": "60140060",
        "codigo_municipio_ibge": 2304400,
        "cidade": "Fortaleza",
        "uf": "CE",
    },
    {
        "match": "unimed",
        "nome": "Unimed Porto Alegre",
        "cnpj": "87096616000196",
        "razao_social": "UNIMED PORTO ALEGRE - COOPERATIVA MEDICA LTDA",
        "email_nf": None,
        "endereco": "AV VENANCIO AIRES, 1040, FARROUPILHA",
        "cep": "90040192",
        "codigo_municipio_ibge": 4314902,
        "cidade": "Porto Alegre",
        "uf": "RS",
    },
    {
        "match": "geap",
        "nome": "GEAP",
        "cnpj": "03658432001820",
        "razao_social": "GEAP AUTOGESTAO EM SAUDE",
        "email_nf": None,
        "endereco": "R LUCIANA DE ABREU, 416, MOINHOS DE VENTO",
        "cep": "90570060",
        "codigo_municipio_ibge": 4314902,
        "cidade": "Porto Alegre",
        "uf": "RS",
    },
    {
        "match": "instituto assistencia",
        "nome": "IPE Saúde",
        "cnpj": "30483455000176",
        "razao_social": "INSTITUTO ASSISTENCIA A SAUDE DOS SERVIDORES PUBLICOS DO RS",
        "email_nf": None,
        "endereco": "AV BORGES DE MEDEIROS, 1945, PRAIA DE BELAS",
        "cep": "90110900",
        "codigo_municipio_ibge": 4314902,
        "cidade": "Porto Alegre",
        "uf": "RS",
    },
    {
        "match": "unimed vitoria",
        "nome": "Unimed Vitória",
        "cnpj": "27578434000120",
        "razao_social": "UNIMED VITORIA COOPERATIVA DE TRABALHO MEDICO",
        "email_nf": "pagamentoscoinr@unimedvx.com.br",
        "endereco": "AV CEZAR HILAL, 700, BENTO FERREIRA",
        "cep": "29050903",
        "codigo_municipio_ibge": 3205309,
        "cidade": "Vitoria",
        "uf": "ES",
    },
    {
        "match": "doctor clin",
        "nome": "Doctor Clin",
        "cnpj": "01387625000110",
        "razao_social": "DOCTOR CLIN OPERADORA DE PLANOS DE SAUDE LTDA",
        "email_nf": "cintia.skonetzky@doctorclin.com.br",
        "endereco": "R SETE DE SETEMBRO, 769, ANDAR 10, CENTRO HISTORICO",
        "cep": "90010190",
        "codigo_municipio_ibge": 4314902,
        "cidade": "Porto Alegre",
        "uf": "RS",
    },
]


def norm_name(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", value.strip().lower())


def rest(method: str, url: str, headers: dict, body: object | None = None) -> tuple[int, object]:
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=120) as res:
            raw = res.read().decode()
            return res.status, json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        raw = e.read().decode()
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            payload = raw
        return e.code, payload


def upsert_integracao(base: str, h: dict) -> None:
    rows = [{"chave": "FOCUSNFE_INSCRICAO_MUNICIPAL", "valor": IM_CB_MOVE}]
    code, result = rest(
        "POST",
        f"{base}/rest/v1/integracao_config?on_conflict=chave",
        {**h, "Prefer": "resolution=merge-duplicates"},
        rows,
    )
    if code >= 400:
        raise RuntimeError(f"integracao_config ({code}): {result}")
    print(f"OK FOCUSNFE_INSCRICAO_MUNICIPAL = {IM_CB_MOVE}")


def load_convenios(base: str, h: dict) -> list[dict]:
    code, rows = rest("GET", f"{base}/rest/v1/convenios?select=*", h)
    if code >= 400:
        raise RuntimeError(f"convenios GET ({code}): {rows}")
    return rows  # type: ignore[return-value]


def upsert_convenio(base: str, h: dict, row: dict, existing_id: str | None) -> None:
    basic = {
        "nome": row["nome"],
        "cnpj": row["cnpj"],
        "razao_social": row["razao_social"],
        "email_nf": row.get("email_nf"),
    }
    extended = {
        "endereco": row.get("endereco"),
        "cep": row.get("cep"),
        "cidade": row.get("cidade"),
        "uf": row.get("uf"),
        "codigo_municipio_ibge": row.get("codigo_municipio_ibge"),
    }

    if existing_id:
        code, result = rest("PATCH", f"{base}/rest/v1/convenios?id=eq.{existing_id}", h, basic)
    else:
        code, result = rest("POST", f"{base}/rest/v1/convenios", h, {**basic, "ativo": True})
    if code >= 400:
        raise RuntimeError(f"convenio {row.get('nome')} ({code}): {result}")

    if existing_id:
        code2, result2 = rest(
            "PATCH",
            f"{base}/rest/v1/convenios?id=eq.{existing_id}",
            h,
            {k: v for k, v in extended.items() if v is not None},
        )
        if code2 >= 400:
            print(f"AVISO: campos de endereço não aplicados em {row['nome']} (migration pendente?)")

    print(f"OK convenio: {row['nome']}")


def sync_convenios(base: str, h: dict) -> None:
    existing = load_convenios(base, h)
    by_norm = {norm_name(c["nome"]): c for c in existing}
    used_ids: set[str] = set()

    for spec in sorted(CONVENIOS, key=lambda s: len(s["match"]), reverse=True):
        match_key = spec["match"]
        found = None
    for c in existing:
        if c["id"] in used_ids:
            continue
        n = norm_name(c["nome"])
        cnpj = (c.get("cnpj") or "").strip()
        if spec.get("cnpj") and cnpj == spec["cnpj"]:
            found = c
            break
        if match_key in n or (match_key == n):
            found = c
            break
        if not found and match_key in by_norm:
            found = by_norm[match_key]

        upsert_convenio(base, h, spec, found["id"] if found else None)
        if found:
            used_ids.add(found["id"])


def first_email(cell: object) -> str | None:
    if not cell or not isinstance(cell, str):
        return None
    for part in re.split(r"[,;]", cell):
        email = part.strip()
        if "@" in email:
            return email
    return None


def load_patient_emails() -> dict[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        import subprocess

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import load_workbook

    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        nome, _, email_cell, *_ = (list(row) + [None, None, None])[:3]
        if not nome:
            continue
        email = first_email(email_cell)
        if email:
            out[norm_name(str(nome))] = email
    return out


def sync_patient_emails(base: str, h: dict) -> None:
    emails = load_patient_emails()
    code, pacientes = rest(
        "GET",
        f"{base}/rest/v1/pacientes?select=id,nome,email&ativo=eq.true",
        h,
    )
    if code >= 400:
        raise RuntimeError(f"pacientes GET ({code}): {pacientes}")

    updated = 0
    skipped = 0
    for p in pacientes:  # type: ignore[union-attr]
        key = norm_name(p["nome"])
        email = emails.get(key)
        if not email:
            continue
        if p.get("email"):
            skipped += 1
            continue
        code2, result = rest(
            "PATCH",
            f"{base}/rest/v1/pacientes?id=eq.{p['id']}",
            h,
            {"email": email},
        )
        if code2 >= 400:
            raise RuntimeError(f"paciente {p['nome']} ({code2}): {result}")
        updated += 1

    print(f"OK pacientes: {updated} e-mails importados, {skipped} já tinham e-mail")


def main() -> None:
    load_app_env()
    base = os.environ["SUPABASE_URL"].rstrip("/")
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    h = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }

    upsert_integracao(base, h)
    sync_convenios(base, h)
    sync_patient_emails(base, h)
    print("Dados docs2 aplicados.")


if __name__ == "__main__":
    main()
