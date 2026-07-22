#!/usr/bin/env python3
"""
Reimporta Relatório Financeiro 2026 (abas) para cobrancas.

  python scripts/import_meses_financeiro.py --dry-run
  python scripts/import_meses_financeiro.py --apply --meses 1-7

Por padrão: substitui cobranças migrado_logjur das competências pedidas e reinsere a partir da planilha.
"""
from __future__ import annotations

import argparse
import datetime
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from load_app_env import load_app_env

MES_NOME = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
}
ABA_MES = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "MARCO": 3,
    "ABRIL": 4, "MAIO": 5, "JUNHO": 6, "JULHO": 7,
    "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}
MES_ABA = {v: k for k, v in ABA_MES.items() if k != "MARCO"}


def norm_nome(n: str) -> str:
    n = n.strip().lower()
    n = unicodedata.normalize("NFD", n)
    return "".join(c for c in n if unicodedata.category(c) != "Mn")


def cell(row, idx, default=""):
    if len(row) <= idx or row[idx] is None:
        return default
    return row[idx]


def parse_valor(v) -> float | None:
    if isinstance(v, (int, float)) and float(v) > 0:
        return float(v)
    s = str(v or "").replace("R$", "").replace(" ", "")
    if not s:
        return None
    # BR: 10.280,00 → 10280.00 ; 10280,00 → 10280.00 ; 10.28 → 10.28
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        f = float(s)
        return f if f > 0 else None
    except Exception:
        return None


def parse_datas_retroativas(sit: str, mes_atual: int, ano_atual: int) -> list[tuple[int, int]]:
    found: list[tuple[int, int]] = []

    def dedup(mes: int, ano: int) -> None:
        if ano > ano_atual or (ano == ano_atual and mes >= mes_atual):
            return
        if ano < 2020 or ano > 2030:
            return
        if (mes, ano) not in found:
            found.append((mes, ano))

    for m in re.finditer(r"\b(0?[1-9]|1[0-2])/?(\d{4})\b", sit):
        dedup(int(m.group(1)), int(m.group(2)))
    for m in re.finditer(r"\b(0?[1-9]|1[0-2])/(\d{2})\b", sit):
        dedup(int(m.group(1)), 2000 + int(m.group(2)))

    meses_pt = {
        "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
        "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
    }
    for m in re.finditer(
        r"\b(janeiro|fevereiro|mar[çc]o|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s+(?:de\s+)?(\d{4})\b",
        sit,
        re.I,
    ):
        key = unicodedata.normalize("NFD", m.group(1).lower())
        key = "".join(c for c in key if unicodedata.category(c) != "Mn")
        if key in meses_pt:
            dedup(meses_pt[key], int(m.group(2)))
    return found


def infer_modelo(sit: str) -> str:
    s = sit.lower()
    if "sharepoint" in s:
        return "sharepoint"
    if "unimed" in s:
        return "unimed"
    return "convencional"


def infer_tipo(sit: str) -> str:
    s = sit.lower()
    if re.search(r"judicial|alvará|alvara|processo", s):
        return "judicial"
    if re.search(r"sharepoint|unimed|ccg|bradesco\s+segu|convênio|convenio", s):
        return "convenio"
    return "particular"


def infer_forma_pgto(sit: str) -> str:
    s = sit.lower()
    if re.search(r"\bboleto\b", s):
        return "boleto"
    if re.search(r"\bpix\b", s):
        return "transferencia"
    if re.search(r"\bdeposit", s):
        return "deposito"
    if re.search(r"judicial|alvará|alvara", s):
        return "alvara_judicial"
    if re.search(r"convenio_direto|convênio direto", s):
        return "convenio_direto"
    return "deposito"


def infer_status(sit: str, tem_retro: bool) -> str:
    s = sit.lower()
    if re.search(r"\bpago\b", s):
        return "pago"
    if re.search(r"atrasad", s) or tem_retro:
        return "atrasado"
    if re.search(r"vai faltar|falta pagar", s):
        return "pendente"
    if "sharepoint" in s:
        return "aguardando_convenio"
    if re.search(r"judicial|alvará|alvara", s):
        return "aguardando_alvara"
    return "pendente"


def infer_vencimento(sit: str, mes: int, ano: int) -> str:
    m = re.search(r"dia\s*0?(\d{1,2})", sit, re.I)
    dia = min(int(m.group(1)), 28) if m else 15
    last = (datetime.date(ano + (1 if mes == 12 else 0), 1 if mes == 12 else mes + 1, 1) - datetime.timedelta(days=1)).day
    d = min(dia, last)
    return f"{ano}-{mes:02d}-{d:02d}"


def infer_regime(plano: str) -> str:
    p = plano.strip().lower()
    return "por_sessao" if ("sessão" in p or "sessao" in p) else "mensalista"


def infer_servico(frequencia: str, mes: int, ano: int) -> str:
    f = frequencia.lower()
    suffix = f"{MES_NOME[mes]}/{ano}"
    if "triplo" in f:
        return f"Plano triplo {suffix}"
    if "duplo" in f:
        return f"Plano duplo {suffix}"
    return f"Fisioterapia Neurológica {suffix}"


def deve_ignorar(row) -> bool:
    nome = str(cell(row, 0) or "").strip()
    plano = str(cell(row, 5) or "").strip()
    sit = str(cell(row, 9) or "").lower().strip()
    if not nome or len(nome) < 3:
        return True
    if nome == "Nome do Paciente":
        return True
    # linhas de título/lixo da planilha
    if "altera" in nome.lower() and "agenda" in nome.lower():
        return True
    if plano == "*****":
        return True
    if "sem cobran" in sit:
        return True
    return False


def levenshtein_sim(a: str, b: str) -> float:
    if a == b:
        return 1.0
    la, lb = len(a), len(b)
    if not la or not lb:
        return 0.0
    dp = [[0] * (lb + 1) for _ in range(la + 1)]
    for i in range(la + 1):
        dp[i][0] = i
    for j in range(lb + 1):
        dp[0][j] = j
    for i in range(1, la + 1):
        for j in range(1, lb + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            dp[i][j] = min(dp[i - 1][j] + 1, dp[i][j - 1] + 1, dp[i - 1][j - 1] + cost)
    return 1 - dp[la][lb] / max(la, lb)


@dataclass
class CobrancaRow:
    paciente_nome: str
    match_id: str | None
    novo_p: bool
    tipo: str
    modelo: str
    regime: str
    servico: str
    competencia_mes: int
    competencia_ano: int
    vencimento: str
    valor: float
    status: str
    forma_pgto: str
    qtd_sessoes: int | None
    frequencia: str | None
    dias_semana: str | None
    obs: str
    is_retroativa: bool
    alertas: list[str] = field(default_factory=list)


class Supa:
    def __init__(self):
        url = (os.environ.get("SUPABASE_URL") or os.environ.get("VITE_SUPABASE_URL") or "").strip().strip("\"'")
        key = (
            os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
            or os.environ.get("SUPABASE_PUBLISHABLE_KEY")
            or os.environ.get("VITE_SUPABASE_PUBLISHABLE_KEY")
            or ""
        ).strip().strip("\"'")
        if not url.startswith("http"):
            raise SystemExit(f"SUPABASE_URL inválida: {url!r}")
        if not key:
            raise SystemExit("Chave Supabase ausente")
        self.url = url.rstrip("/")
        self.key = key
        print(f"Supabase: {self.url}")

    def _req(self, method: str, path: str, body=None, prefer: str | None = None, extra_headers: dict | None = None):
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
        }
        if prefer:
            headers["Prefer"] = prefer
        if extra_headers:
            headers.update(extra_headers)
        data = None if body is None else json.dumps(body).encode()
        # Keep PostgREST operators; encode spaces and other unsafe chars in filters.
        safe_path = urllib.parse.quote(path, safe="/?&=(),.*:_-")
        req = urllib.request.Request(f"{self.url}/rest/v1/{safe_path}", data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=120) as res:
                raw = res.read().decode()
                return json.loads(raw) if raw else None, res.headers
        except urllib.error.HTTPError as e:
            detail = e.read().decode()[:800]
            raise RuntimeError(f"HTTP {e.code} {path}: {detail}") from e

    def get_pacientes(self):
        rows = []
        offset = 0
        while True:
            chunk, _ = self._req("GET", f"pacientes?select=id,nome,tipo&order=nome&offset={offset}&limit=1000")
            if not chunk:
                break
            rows.extend(chunk)
            if len(chunk) < 1000:
                break
            offset += 1000
        return rows

    def delete_migradas(self, ano: int, meses: list[int]) -> int:
        """Apaga cobranças migradas das competências (inclui retroativas migrado_logjur)."""
        mes_or = ",".join(str(m) for m in meses)
        # PostgREST: or=(observacoes.ilike.*migrado_logjur*,observacoes.ilike.*Retroativa*)
        filt = (
            f"competencia_ano=eq.{ano}&competencia_mes=in.({mes_or})"
            f"&or=(observacoes.ilike.*migrado_logjur*,observacoes.ilike.*Retroativa*)"
        )
        # count first
        _, headers = self._req(
            "GET",
            f"cobrancas?select=id&{filt}",
            prefer="count=exact",
            extra_headers={"Range": "0-0"},
        )
        cr = headers.get("Content-Range") or headers.get("content-range") or ""
        # Content-Range: 0-0/123
        total = 0
        if "/" in cr:
            try:
                total = int(cr.split("/")[-1])
            except Exception:
                total = 0
        self._req("DELETE", f"cobrancas?{filt}", prefer="return=minimal")
        return total

    def insert_paciente(self, row: dict) -> str:
        data, _ = self._req("POST", "pacientes", row, prefer="return=representation")
        return data[0]["id"]

    def insert_cobranca(self, row: dict) -> None:
        self._req("POST", "cobrancas", row, prefer="return=minimal")


def resolve_aba_name(wb_names: list[str], mes: int) -> str | None:
    wanted = MES_ABA.get(mes)
    for n in wb_names:
        if ABA_MES.get(n.upper()) == mes:
            return n
        if wanted and n.upper().replace("Ç", "C") == wanted.replace("Ç", "C"):
            return n
    return None


def build_rows(xlsx: Path, pacientes_db: list[dict], meses: list[int], ano: int = 2026) -> tuple[list[CobrancaRow], dict[int, int]]:
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    cobrancas: list[CobrancaRow] = []
    vazios_por_mes: dict[int, int] = {m: 0 for m in meses}
    novos: set[str] = set()

    for mes in meses:
        aba = resolve_aba_name(wb.sheetnames, mes)
        if not aba:
            print(f"  ! Aba do mes {mes} nao encontrada")
            continue
        rows = list(wb[aba].iter_rows(values_only=True))
        dados = rows[2:]
        print(f"  Aba {aba}: {len(dados)} linhas brutas")

        for row in dados:
            if not row or deve_ignorar(row):
                continue
            nome = str(cell(row, 0) or "").strip()
            frequencia = str(cell(row, 2) or "").strip()
            dias = str(cell(row, 3) or "").strip()
            plano = str(cell(row, 5) or "").strip()
            valor = parse_valor(cell(row, 7, None))
            sit = str(cell(row, 9) or "").strip()
            qtd_raw = cell(row, 4, None)

            if valor is None:
                vazios_por_mes[mes] += 1
                continue

            nom_n = norm_nome(nome)
            match = None
            for p in pacientes_db:
                pn = norm_nome(p["nome"])
                if pn == nom_n or pn in nom_n or nom_n in pn or levenshtein_sim(nom_n, pn) >= 0.82:
                    match = p
                    break

            tipo = infer_tipo(sit)
            modelo = infer_modelo(sit)
            retro = parse_datas_retroativas(sit, mes, ano)
            # só gera retroativa se o mês destino também está no lote (evita lixo fora do range)
            # Na verdade R5 deve criar em qualquer mês passado — manter todas
            tem_retro = len(retro) > 0
            total_comp = len(retro) + 1
            valor_por = round(valor / total_comp, 2)
            valor_ult = round(valor - valor_por * (total_comp - 1), 2)

            try:
                qtd = int(qtd_raw) if qtd_raw not in (None, "") else None
            except Exception:
                qtd = None

            base_obs = f"migrado_logjur | {sit}".strip()
            alertas = []
            if not match and nom_n not in novos:
                alertas.append("novo paciente")

            for i, (rm, ra) in enumerate(retro):
                is_last = i == len(retro) - 1 and len(retro) == total_comp - 1
                cobrancas.append(
                    CobrancaRow(
                        paciente_nome=nome,
                        match_id=match["id"] if match else None,
                        novo_p=not match,
                        tipo=tipo,
                        modelo=modelo,
                        regime=infer_regime(plano),
                        servico=f"{infer_servico(frequencia, rm, ra)} [retroativa]",
                        competencia_mes=rm,
                        competencia_ano=ra,
                        vencimento=infer_vencimento(sit, rm, ra),
                        valor=valor_ult if is_last else valor_por,
                        status="regularizar_retroativa",
                        forma_pgto=infer_forma_pgto(sit),
                        qtd_sessoes=None,
                        frequencia=frequencia or None,
                        dias_semana=dias or None,
                        obs=f"Retroativa detectada no relatório financeiro | {base_obs}",
                        is_retroativa=True,
                    )
                )

            cobrancas.append(
                CobrancaRow(
                    paciente_nome=nome,
                    match_id=match["id"] if match else None,
                    novo_p=not match,
                    tipo=tipo,
                    modelo=modelo,
                    regime=infer_regime(plano),
                    servico=infer_servico(frequencia, mes, ano),
                    competencia_mes=mes,
                    competencia_ano=ano,
                    vencimento=infer_vencimento(sit, mes, ano),
                    valor=valor_ult if tem_retro else valor,
                    status=infer_status(sit, tem_retro),
                    forma_pgto=infer_forma_pgto(sit),
                    qtd_sessoes=qtd,
                    frequencia=frequencia or None,
                    dias_semana=dias or None,
                    obs=base_obs,
                    is_retroativa=False,
                    alertas=alertas,
                )
            )
            if not match:
                novos.add(nom_n)

    wb.close()
    return cobrancas, vazios_por_mes


def parse_meses(s: str) -> list[int]:
    s = s.strip()
    if "-" in s:
        a, b = s.split("-", 1)
        return list(range(int(a), int(b) + 1))
    return [int(x) for x in s.split(",") if x.strip()]


def main() -> None:
    load_app_env()
    # limpa URL quoted do shell
    for k in ("SUPABASE_URL", "VITE_SUPABASE_URL"):
        if k in os.environ:
            os.environ[k] = os.environ[k].strip().strip("\"'")

    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default="scripts/drive_import/relatorio_financeiro_2026_fresh.xlsx")
    ap.add_argument("--meses", default="1-7", help="ex: 1-6 ou 1,2,3,7")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--ano", type=int, default=2026)
    args = ap.parse_args()

    xlsx = Path(args.file)
    if not xlsx.exists():
        raise SystemExit(f"Arquivo nao encontrado: {xlsx}")

    meses = parse_meses(args.meses)
    sb = Supa()
    pacientes = sb.get_pacientes()
    print(f"Arquivo: {xlsx}")
    print(f"Meses: {meses} | Ano: {args.ano} | Modo: {'APPLY' if args.apply else 'DRY-RUN'}")
    print(f"Pacientes DB: {len(pacientes)}")

    cobrancas, vazios = build_rows(xlsx, pacientes, meses, args.ano)
    por_mes: dict[int, list[CobrancaRow]] = {m: [] for m in meses}
    for c in cobrancas:
        if c.competencia_ano == args.ano and c.competencia_mes in por_mes:
            por_mes[c.competencia_mes].append(c)

    print("\nResumo a inserir (competencia do mes; retroativas de outros meses entram no lote total):")
    for m in meses:
        lst = por_mes[m]
        soma = sum(c.valor for c in lst)
        print(f"  {MES_NOME[m]}/{args.ano}: {len(lst)} cobrancas | soma R$ {soma:,.2f} | vazios planilha={vazios.get(m,0)}")
    print(f"Total registros (incl. retro p/ fora do range): {len(cobrancas)}")
    print(f"Pacientes novos: {len({norm_nome(c.paciente_nome) for c in cobrancas if c.novo_p})}")

    if not args.apply:
        print("\nDRY-RUN ok. Use --apply para apagar migradas e inserir.")
        return

    # competências a limpar: meses do lote + destinos de retroativas dentro de 2026
    limpar = sorted(set(meses) | {c.competencia_mes for c in cobrancas if c.competencia_ano == args.ano})
    print(f"\nApagando migrado_logjur/retroativa em {args.ano} meses {limpar}...")
    deleted = sb.delete_migradas(args.ano, limpar)
    print(f"  removidas (estimativa Content-Range): {deleted}")

    cache: dict[str, str] = {}
    ok = 0
    errs: list[str] = []
    for c in cobrancas:
        nn = norm_nome(c.paciente_nome)
        pac_id = c.match_id or cache.get(nn)
        if not pac_id:
            try:
                pac_id = sb.insert_paciente(
                    {
                        "nome": c.paciente_nome,
                        "tipo": c.tipo,
                        "modelo_relatorio_preferido": c.modelo,
                        "frequencia_atendimento": c.frequencia,
                        "dias_semana": c.dias_semana,
                    }
                )
                cache[nn] = pac_id
                print(f"  + paciente {c.paciente_nome}")
            except Exception as e:
                errs.append(f"PAC {c.paciente_nome}: {e}")
                continue
        try:
            sb.insert_cobranca(
                {
                    "paciente_id": pac_id,
                    "competencia_mes": c.competencia_mes,
                    "competencia_ano": c.competencia_ano,
                    "tipo": c.tipo,
                    "regime": c.regime,
                    "servico": c.servico,
                    "valor": c.valor,
                    "forma_pagamento": c.forma_pgto,
                    "vencimento": c.vencimento,
                    "status": c.status,
                    "qtd_sessoes": c.qtd_sessoes,
                    "frequencia_atendimento": c.frequencia,
                    "dias_semana": c.dias_semana,
                    "observacoes": c.obs,
                }
            )
            ok += 1
        except Exception as e:
            errs.append(f"COB {c.paciente_nome} {c.competencia_mes}/{c.competencia_ano}: {e}")

    print(f"\nInseridas: {ok} | Erros: {len(errs)}")
    for e in errs[:30]:
        print(" ", e)


if __name__ == "__main__":
    main()
