#!/usr/bin/env python3
"""Importa aba JULHO do Relatório Financeiro 2026 para cobrancas (espelha import-relatorio-financeiro.ts)."""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from load_app_env import load_app_env
from lib.retroativos_valor import calc_valor_mes_atual, calc_valor_retroativo

MES_NOME = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
}


def norm_nome(n: str) -> str:
    n = n.strip().lower()
    n = unicodedata.normalize("NFD", n)
    return "".join(c for c in n if unicodedata.category(c) != "Mn")


def parse_valor(v) -> float | None:
    if isinstance(v, (int, float)) and float(v) > 0:
        return float(v)
    s = str(v or "").replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        f = float(s)
        return f if f > 0 else None
    except Exception:
        return None


def parse_datas_retroativas(sit: str, mes_atual: int, ano_atual: int) -> list[tuple[int, int]]:
    found: list[tuple[int, int]] = []

    def dedup(mes: int, ano: int) -> None:
        if ano > ano_atual or (ano == ano_atual and mes >= mes_atual):
            return
        if ano < 2020 or ano > 2030:
            return
        if (mes, ano) not in found:
            found.append((mes, ano))

    for m in re.finditer(r"\b(0?[1-9]|1[0-2])/?(\d{4})\b", sit):
        dedup(int(m.group(1)), int(m.group(2)))
    for m in re.finditer(r"\b(0?[1-9]|1[0-2])/(\d{2})\b", sit):
        dedup(int(m.group(1)), 2000 + int(m.group(2)))

    meses_pt = {
        "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
        "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
    }
    for m in re.finditer(
        r"\b(janeiro|fevereiro|mar[çc]o|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s+(?:de\s+)?(\d{4})\b",
        sit,
        re.I,
    ):
        key = unicodedata.normalize("NFD", m.group(1).lower())
        key = "".join(c for c in key if unicodedata.category(c) != "Mn")
        if key in meses_pt:
            dedup(meses_pt[key], int(m.group(2)))
    return found


def infer_modelo(sit: str) -> str:
    s = sit.lower()
    if "sharepoint" in s:
        return "sharepoint"
    if "unimed" in s:
        return "unimed"
    return "convencional"


def infer_tipo(sit: str) -> str:
    s = sit.lower()
    if re.search(r"judicial|alvará|alvara|processo", s):
        return "judicial"
    if re.search(r"sharepoint|unimed|ccg|bradesco\s+segu|convênio|convenio", s):
        return "convenio"
    return "particular"


def infer_forma_pgto(sit: str) -> str:
    s = sit.lower()
    if re.search(r"\bboleto\b", s):
        return "boleto"
    if re.search(r"\bpix\b", s):
        return "transferencia"
    if re.search(r"\bdeposit", s):
        return "deposito"
    if re.search(r"judicial|alvará|alvara", s):
        return "alvara_judicial"
    if re.search(r"convenio_direto|convênio direto", s):
        return "convenio_direto"
    return "deposito"


def infer_status(sit: str, tem_retro: bool) -> str:
    s = sit.lower()
    if re.search(r"\bpago\b", s):
        return "pago"
    if re.search(r"atrasad", s) or tem_retro:
        return "atrasado"
    if re.search(r"vai faltar|falta pagar", s):
        return "pendente"
    if "sharepoint" in s:
        return "aguardando_convenio"
    if re.search(r"judicial|alvará|alvara", s):
        return "aguardando_alvara"
    return "pendente"


def infer_vencimento(sit: str, mes: int, ano: int) -> str:
    m = re.search(r"dia\s*0?(\d{1,2})", sit, re.I)
    dia = min(int(m.group(1)), 28) if m else 15
    # last day of month
    if mes == 12:
        last = 31
    else:
        import datetime
        last = (datetime.date(ano, mes + 1, 1) - datetime.timedelta(days=1)).day
    d = min(dia, last)
    return f"{ano}-{mes:02d}-{d:02d}"


def infer_regime(plano: str) -> str:
    p = plano.strip().lower()
    return "por_sessao" if ("sessão" in p or "sessao" in p) else "mensalista"


def infer_servico(frequencia: str, mes: int, ano: int) -> str:
    f = frequencia.lower()
    suffix = f"{MES_NOME[mes]}/{ano}"
    if "triplo" in f:
        return f"Plano triplo {suffix}"
    if "duplo" in f:
        return f"Plano duplo {suffix}"
    return f"Fisioterapia Neurológica {suffix}"


def cell(row, idx, default=""):
    if len(row) <= idx or row[idx] is None:
        return default
    return row[idx]


def deve_ignorar(row) -> bool:
    nome = str(cell(row, 0) or "").strip()
    plano = str(cell(row, 5) or "").strip()
    sit = str(cell(row, 9) or "").lower().strip()
    if not nome or len(nome) < 3:
        return True
    if nome == "Nome do Paciente":
        return True
    if plano == "*****":
        return True
    if "sem cobran" in sit:
        return True
    return False


def levenshtein_sim(a: str, b: str) -> float:
    if a == b:
        return 1.0
    la, lb = len(a), len(b)
    if not la or not lb:
        return 0.0
    dp = [[0] * (lb + 1) for _ in range(la + 1)]
    for i in range(la + 1):
        dp[i][0] = i
    for j in range(lb + 1):
        dp[0][j] = j
    for i in range(1, la + 1):
        for j in range(1, lb + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            dp[i][j] = min(dp[i - 1][j] + 1, dp[i][j - 1] + 1, dp[i - 1][j - 1] + cost)
    return 1 - dp[la][lb] / max(la, lb)


@dataclass
class CobrancaRow:
    paciente_nome: str
    match_id: str | None
    novo_p: bool
    tipo: str
    modelo: str
    regime: str
    servico: str
    competencia_mes: int
    competencia_ano: int
    vencimento: str
    valor: float
    status: str
    forma_pgto: str
    qtd_sessoes: int | None
    frequencia: str | None
    dias_semana: str | None
    obs: str
    is_retroativa: bool
    alertas: list[str] = field(default_factory=list)


class Supa:
    def __init__(self):
        url = (os.environ.get("SUPABASE_URL") or os.environ.get("VITE_SUPABASE_URL") or "").strip().strip("\"'")
        key = (
            os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
            or os.environ.get("SUPABASE_PUBLISHABLE_KEY")
            or os.environ.get("VITE_SUPABASE_PUBLISHABLE_KEY")
            or ""
        ).strip().strip("\"'")
        if not url.startswith("http"):
            raise SystemExit(f"SUPABASE_URL inválida: {url!r}")
        if not key:
            raise SystemExit("SUPABASE_SERVICE_ROLE_KEY / PUBLISHABLE_KEY ausente")
        self.url = url.rstrip("/")
        self.key = key
        print(f"Supabase: {self.url}")

    def _req(self, method: str, path: str, body=None, prefer: str | None = None):
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
        }
        if prefer:
            headers["Prefer"] = prefer
        data = None if body is None else json.dumps(body).encode()
        req = urllib.request.Request(f"{self.url}/rest/v1/{path}", data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=60) as res:
                raw = res.read().decode()
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as e:
            detail = e.read().decode()[:500]
            raise RuntimeError(f"HTTP {e.code} {path}: {detail}") from e

    def get_pacientes(self):
        rows = []
        offset = 0
        while True:
            chunk = self._req(
                "GET",
                f"pacientes?select=id,nome,tipo,valor_mensal&order=nome&offset={offset}&limit=1000",
            )
            if not chunk:
                break
            rows.extend(chunk)
            if len(chunk) < 1000:
                break
            offset += 1000
        return rows

    def insert_paciente(self, row: dict) -> str:
        data = self._req("POST", "pacientes", row, prefer="return=representation")
        return data[0]["id"]

    def insert_cobranca(self, row: dict) -> None:
        self._req("POST", "cobrancas", row, prefer="return=minimal")


def build_rows(xlsx: Path, pacientes_db: list[dict]) -> tuple[list[CobrancaRow], list[str]]:
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    if "JULHO" not in wb.sheetnames:
        raise SystemExit(f"Aba JULHO não encontrada. Abas: {wb.sheetnames}")
    rows = list(wb["JULHO"].iter_rows(values_only=True))
    wb.close()
    dados = rows[2:]
    mes, ano = 7, 2026
    cobrancas: list[CobrancaRow] = []
    vazios: list[str] = []
    novos = set()

    for row in dados:
        if not row or deve_ignorar(row):
            continue
        nome = str(cell(row, 0) or "").strip()
        frequencia = str(cell(row, 2) or "").strip()
        dias = str(cell(row, 3) or "").strip()
        plano = str(cell(row, 5) or "").strip()
        valor = parse_valor(cell(row, 7, None))
        sit = str(cell(row, 9) or "").strip()
        qtd_raw = cell(row, 4, None)

        if valor is None:
            vazios.append(nome)
            continue

        nom_n = norm_nome(nome)
        match = None
        for p in pacientes_db:
            pn = norm_nome(p["nome"])
            if pn == nom_n or pn in nom_n or nom_n in pn or levenshtein_sim(nom_n, pn) >= 0.82:
                match = p
                break

        tipo = infer_tipo(sit)
        modelo = infer_modelo(sit)
        retro = parse_datas_retroativas(sit, mes, ano)
        tem_retro = len(retro) > 0
        valor_mensal_pac = float(match["valor_mensal"]) if match and match.get("valor_mensal") else None
        valor_retro = calc_valor_retroativo(valor_mensal_pac, valor)
        valor_mes = calc_valor_mes_atual(valor)

        try:
            qtd = int(qtd_raw) if qtd_raw not in (None, "") else None
        except Exception:
            qtd = None

        base_obs = f"migrado_logjur | {sit}".strip()
        alertas = []
        if not match and nom_n not in novos:
            alertas.append("novo paciente")
        if valor > 50000:
            alertas.append(f"VALOR ALTO: R$ {valor}")
        if tem_retro:
            alertas.append(f"{len(retro)} retroativa(s)")

        for i, (rm, ra) in enumerate(retro):
            cobrancas.append(
                CobrancaRow(
                    paciente_nome=nome,
                    match_id=match["id"] if match else None,
                    novo_p=not match,
                    tipo=tipo,
                    modelo=modelo,
                    regime=infer_regime(plano),
                    servico=f"{infer_servico(frequencia, rm, ra)} [retroativa]",
                    competencia_mes=rm,
                    competencia_ano=ra,
                    vencimento=infer_vencimento(sit, rm, ra),
                    valor=valor_retro,
                    status="regularizar_retroativa",
                    forma_pgto=infer_forma_pgto(sit),
                    qtd_sessoes=None,
                    frequencia=frequencia or None,
                    dias_semana=dias or None,
                    obs=f"Retroativa detectada no relatório financeiro | {base_obs}",
                    is_retroativa=True,
                )
            )

        cobrancas.append(
            CobrancaRow(
                paciente_nome=nome,
                match_id=match["id"] if match else None,
                novo_p=not match,
                tipo=tipo,
                modelo=modelo,
                regime=infer_regime(plano),
                servico=infer_servico(frequencia, mes, ano),
                competencia_mes=mes,
                competencia_ano=ano,
                vencimento=infer_vencimento(sit, mes, ano),
                valor=valor_mes,
                status=infer_status(sit, tem_retro),
                forma_pgto=infer_forma_pgto(sit),
                qtd_sessoes=qtd,
                frequencia=frequencia or None,
                dias_semana=dias or None,
                obs=base_obs,
                is_retroativa=False,
                alertas=alertas,
            )
        )
        if not match:
            novos.add(nom_n)

    return cobrancas, vazios


def main() -> None:
    load_app_env()
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--file",
        default=str(Path("scripts/drive_import/relatorio_financeiro_2026_fresh.xlsx")),
    )
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    xlsx = Path(args.file)
    if not xlsx.exists():
        raise SystemExit(f"Arquivo não encontrado: {xlsx}")

    sb = Supa()
    pacientes = sb.get_pacientes()
    cobrancas, vazios = build_rows(xlsx, pacientes)

    jul = [c for c in cobrancas if c.competencia_mes == 7]
    retro = [c for c in cobrancas if c.is_retroativa]
    novos = {norm_nome(c.paciente_nome) for c in cobrancas if c.novo_p}
    soma_jul = sum(c.valor for c in jul)

    print(f"Arquivo: {xlsx}")
    print(f"Modo: {'APPLY' if args.apply else 'DRY-RUN'}")
    print(f"Pacientes DB: {len(pacientes)}")
    print(f"Linhas valor vazio (ignoradas): {len(vazios)}")
    print(f"Cobranças a inserir: {len(cobrancas)} (julho={len(jul)}, retro={len(retro)})")
    print(f"Pacientes novos: {len(novos)}")
    print(f"Soma julho: R$ {soma_jul:,.2f}")
    print("\nAmostra julho (20):")
    for c in jul[:20]:
        print(f"  {c.paciente_nome[:35]:35} R$ {c.valor:10.2f} {c.status:22} {c.forma_pgto}")

    if not args.apply:
        print("\nDRY-RUN ok. Rode com --apply para inserir.")
        return

    cache: dict[str, str] = {}
    ok = 0
    errs: list[str] = []
    for c in cobrancas:
        nn = norm_nome(c.paciente_nome)
        pac_id = c.match_id or cache.get(nn)
        if not pac_id:
            try:
                pac_id = sb.insert_paciente(
                    {
                        "nome": c.paciente_nome,
                        "tipo": c.tipo,
                        "modelo_relatorio_preferido": c.modelo,
                        "frequencia_atendimento": c.frequencia,
                        "dias_semana": c.dias_semana,
                    }
                )
                cache[nn] = pac_id
                print(f"  + paciente {c.paciente_nome}")
            except Exception as e:
                errs.append(f"PAC {c.paciente_nome}: {e}")
                continue
        try:
            sb.insert_cobranca(
                {
                    "paciente_id": pac_id,
                    "competencia_mes": c.competencia_mes,
                    "competencia_ano": c.competencia_ano,
                    "tipo": c.tipo,
                    "regime": c.regime,
                    "servico": c.servico,
                    "valor": c.valor,
                    "forma_pagamento": c.forma_pgto,
                    "vencimento": c.vencimento,
                    "status": c.status,
                    "qtd_sessoes": c.qtd_sessoes,
                    "frequencia_atendimento": c.frequencia,
                    "dias_semana": c.dias_semana,
                    "observacoes": c.obs,
                }
            )
            ok += 1
        except Exception as e:
            errs.append(f"COB {c.paciente_nome} {c.competencia_mes}/{c.competencia_ano}: {e}")

    print(f"\nInseridas: {ok} | Erros: {len(errs)}")
    for e in errs[:20]:
        print(" ", e)


if __name__ == "__main__":
    main()
