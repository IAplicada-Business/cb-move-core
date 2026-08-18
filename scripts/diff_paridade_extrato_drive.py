#!/usr/bin/env python3
"""Fase 2b — diff célula a célula: planilha master Drive × cobranças no sistema.

Uso:
  python3 scripts/diff_paridade_extrato_drive.py
  python3 scripts/diff_paridade_extrato_drive.py --mes JULHO --amostra

Requer: openpyxl, e opcionalmente SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY em .env / .env.app
para comparar com o banco. Sem credenciais, só audita a planilha (headers + amostras).
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "scripts" / "drive_import" / "relatorio_financeiro_2026.xlsx"
OUT = ROOT / "scripts" / "out" / "fase2-paridade"
OUT.mkdir(parents=True, exist_ok=True)

# Colunas canônicas do extrato CB MOVE (src/lib/domain/extrato-financeiro.ts)
CANON = [
    "Nome do Paciente",
    "Avaliação",
    "Frequência",
    "Dias da Semana",
    "Nº Sessões",
    "Plano",
    "R$ Sessão/Mês",
    "R$ Previsto",
    "R$ Recebido",
    "SITUAÇÃO",
]

MES_NUM = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARÇO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}

AMOSTRAS = ["Airton Tonelo", "Amanda Pavan", "Kayhan"]


def norm_header(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).strip()
    aliases = {
        "Previsto": "R$ Previsto",
        "R$ Referente": "R$ Previsto",  # drift AGOSTO
    }
    return aliases.get(s, s)


def money(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower()
    if s in {"pago", "pagou", "-"}:
        return None
    s = s.replace("r$", "").replace(".", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s) if s else None
    except ValueError:
        return None


def load_sheet_rows(ws) -> tuple[list[str], list[dict[str, Any]]]:
    headers_raw = [ws.cell(2, c).value for c in range(1, 15)]
    headers = []
    for h in headers_raw:
        if h is None and headers:
            break
        if h is None:
            continue
        headers.append(norm_header(h))

    rows: list[dict[str, Any]] = []
    for r in range(3, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        if not nome or len(str(nome).strip()) < 3:
            continue
        nome_s = str(nome).strip()
        if nome_s.lower() == "nome do paciente":
            continue
        row = {"_row": r}
        for i, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, i).value
        plano = str(row.get("Plano") or "").strip()
        if plano in {"*****", "-"} and not money(row.get("R$ Previsto")):
            continue
        sit = str(row.get("SITUAÇÃO") or "").strip().lower()
        if sit == "sem cobrança":
            continue
        rows.append(row)
    return headers, rows


@dataclass
class CellDiff:
    paciente: str
    campo: str
    drive: Any
    sistema: Any
    status: str  # ok | diverge | ausente_drive | ausente_sistema


def load_env() -> None:
    for name in (".env.app", ".env"):
        path = ROOT / name
        if not path.is_file():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip("\"'"))


def fetch_sistema(mes: int, ano: int = 2026, sistema_json: Path | None = None) -> list[dict[str, Any]]:
    if sistema_json and sistema_json.is_file():
        data = json.loads(sistema_json.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else data.get("rows") or data.get("data") or []

    load_env()
    url = os.environ.get("SUPABASE_URL") or "https://grlkbtnwvxorlfglyzid.supabase.co"
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_SECRET_KEY")
    if not key:
        return []

    import urllib.parse
    import urllib.request

    params = urllib.parse.urlencode(
        {
            "select": "id,valor,status,regime,qtd_sessoes,frequencia_atendimento,dias_semana,observacoes,pago_em,pacientes(nome,valor_mensal,valor_sessao,regime_cobranca,frequencia_atendimento,dias_semana,criado_em)",
            "competencia_mes": f"eq.{mes}",
            "competencia_ano": f"eq.{ano}",
            "status": "neq.cancelado",
            "order": "created_at.asc",
        }
    )
    req = urllib.request.Request(
        f"{url}/rest/v1/cobrancas?{params}",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode())


def sistema_to_row(c: dict[str, Any]) -> dict[str, Any]:
    # Aceita formato nested (REST) ou flat (SQL export)
    p = c.get("pacientes") or {}
    if not p and c.get("paciente_nome"):
        p = {
            "nome": c.get("paciente_nome"),
            "valor_mensal": c.get("valor_mensal"),
            "valor_sessao": c.get("valor_sessao"),
            "regime_cobranca": c.get("regime_cobranca") or c.get("p_regime"),
            "frequencia_atendimento": c.get("p_freq") or c.get("frequencia_atendimento"),
            "dias_semana": c.get("p_dias") or c.get("dias_semana"),
            "criado_em": c.get("criado_em"),
        }
    regime = c.get("regime") or p.get("regime_cobranca")
    plano = "Por Sessão" if regime == "por_sessao" else ("Mensalista" if regime == "mensalista" else "—")
    valor = float(c.get("valor") or 0)
    pago = c.get("status") == "pago"
    obs = c.get("observacoes") or ""
    sit = obs
    m = re.search(r"migrado_logjur\s*\|\s*(.+)", obs, re.I)
    if m:
        sit = m.group(1).strip()
    elif not str(sit).strip() or str(sit).strip() == "migrado_logjur |":
        sit = "PAGO" if pago else (c.get("status") or "")

    unit = p.get("valor_mensal") if regime == "mensalista" else p.get("valor_sessao")
    if unit is None:
        unit = p.get("valor_mensal") or p.get("valor_sessao")

    return {
        "Nome do Paciente": p.get("nome") or "—",
        "Frequência": c.get("frequencia_atendimento") or p.get("frequencia_atendimento"),
        "Dias da Semana": c.get("dias_semana") or p.get("dias_semana"),
        "Nº Sessões": c.get("qtd_sessoes"),
        "Plano": plano,
        "R$ Sessão/Mês": float(unit) if unit is not None else None,
        "R$ Previsto": valor,
        "R$ Recebido": valor if pago else None,
        "SITUAÇÃO": sit,
    }


def approx_name(a: str, b: str) -> bool:
    na = re.sub(r"\s+", " ", a.lower().strip())
    nb = re.sub(r"\s+", " ", b.lower().strip())
    if na == nb:
        return True
    if na in nb or nb in na:
        return True
    # primeiro + último token
    ta, tb = na.split(), nb.split()
    if ta and tb and ta[0] == tb[0] and ta[-1] == tb[-1]:
        return True
    return False


def cmp_money(a: Any, b: Any) -> bool:
    ma, mb = money(a), money(b)
    if ma is None and mb is None:
        return True
    if ma is None or mb is None:
        return False
    return abs(ma - mb) <= 0.05


def cmp_text(a: Any, b: Any) -> bool:
    sa = "" if a is None else str(a).strip().lower()
    sb = "" if b is None else str(b).strip().lower()
    sa = re.sub(r"\s+", " ", sa)
    sb = re.sub(r"\s+", " ", sb)
    return sa == sb or (not sa and not sb)


def diff_paciente(drive: dict[str, Any], sistema: dict[str, Any] | None) -> list[CellDiff]:
    nome = str(drive.get("Nome do Paciente") or "")
    out: list[CellDiff] = []
    campos = [
        "Frequência",
        "Dias da Semana",
        "Nº Sessões",
        "Plano",
        "R$ Sessão/Mês",
        "R$ Previsto",
        "R$ Recebido",
        "SITUAÇÃO",
    ]
    if sistema is None:
        for c in campos:
            out.append(CellDiff(nome, c, drive.get(c), None, "ausente_sistema"))
        return out

    for c in campos:
        dv, sv = drive.get(c), sistema.get(c)
        if c.startswith("R$") or c == "Nº Sessões":
            ok = cmp_money(dv, sv) if c.startswith("R$") else (
                (dv is None and sv is None)
                or (dv is not None and sv is not None and abs(float(dv) - float(sv)) < 0.1)
            )
        else:
            ok = cmp_text(dv, sv)
        out.append(
            CellDiff(
                nome,
                c,
                dv if not isinstance(dv, datetime) else dv.isoformat(),
                sv,
                "ok" if ok else "diverge",
            )
        )
    return out


def analyze_headers(headers: list[str]) -> dict[str, Any]:
    missing = [c for c in CANON if c not in headers]
    extra = [h for h in headers if h and h not in CANON]
    return {
        "headers_drive": headers,
        "canon": CANON,
        "missing_vs_canon": missing,
        "extra_vs_canon": extra,
        "header_parity": "ok" if not missing else "diverge",
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--mes", default="JULHO", help="Nome da aba (ex: JULHO)")
    ap.add_argument("--ano", type=int, default=2026)
    ap.add_argument("--amostra", action="store_true", help="Só pacientes amostra (Airton/Amanda/Kayhan)")
    ap.add_argument(
        "--sistema-json",
        type=Path,
        help="JSON com cobranças do sistema (lista flat ou REST nested)",
    )
    args = ap.parse_args()

    if not XLSX.exists():
        print(f"Baixe a planilha: python3 scripts/download_drive_files.py\nFalta: {XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    sheet = args.mes.upper()
    if sheet not in wb.sheetnames:
        print(f"Aba {sheet} não existe. Disponíveis: {wb.sheetnames}", file=sys.stderr)
        return 1

    headers, drive_rows = load_sheet_rows(wb[sheet])
    header_report = analyze_headers(headers)

    mes_num = MES_NUM[sheet]
    sistema_raw = fetch_sistema(mes_num, args.ano, args.sistema_json)
    sistema_rows = [sistema_to_row(c) for c in sistema_raw]

    if args.amostra:
        drive_rows = [
            r
            for r in drive_rows
            if any(a.lower() in str(r.get("Nome do Paciente") or "").lower() for a in AMOSTRAS)
        ]

    diffs: list[CellDiff] = []
    matched_sys: set[int] = set()

    for d in drive_rows:
        nome = str(d.get("Nome do Paciente") or "")
        sys_match = None
        for i, s in enumerate(sistema_rows):
            if i in matched_sys:
                continue
            if approx_name(nome, str(s.get("Nome do Paciente") or "")):
                sys_match = s
                matched_sys.add(i)
                break
        diffs.extend(diff_paciente(d, sys_match))

    only_sistema = []
    for i, s in enumerate(sistema_rows):
        if i in matched_sys:
            continue
        nome = str(s.get("Nome do Paciente") or "")
        if args.amostra and not any(a.lower() in nome.lower() for a in AMOSTRAS):
            continue
        only_sistema.append(nome)
        for c in ["Plano", "R$ Previsto", "R$ Recebido", "SITUAÇÃO"]:
            diffs.append(CellDiff(nome, c, None, s.get(c), "ausente_drive"))

    summary = {
        "gerado_em": datetime.utcnow().isoformat() + "Z",
        "aba": sheet,
        "ano": args.ano,
        "header": header_report,
        "drive_linhas": len(drive_rows),
        "sistema_linhas": len(sistema_rows) if sistema_raw else None,
        "sistema_disponivel": bool(sistema_raw),
        "contagem": {
            "ok": sum(1 for d in diffs if d.status == "ok"),
            "diverge": sum(1 for d in diffs if d.status == "diverge"),
            "ausente_sistema": sum(1 for d in diffs if d.status == "ausente_sistema"),
            "ausente_drive": sum(1 for d in diffs if d.status == "ausente_drive"),
        },
        "somente_sistema": only_sistema[:50],
        "diffs": [asdict(d) for d in diffs if d.status != "ok"],
    }

    # all sheets header audit
    all_headers = {}
    for name in wb.sheetnames:
        h, _ = load_sheet_rows(wb[name])
        all_headers[name] = analyze_headers(h)

    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    out_json = OUT / f"diff-{sheet.lower()}-{stamp}.json"
    out_md = OUT / f"diff-{sheet.lower()}-{stamp}.md"
    out_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    lines = [
        f"# Diff paridade — {sheet}/{args.ano}",
        "",
        f"Gerado: `{summary['gerado_em']}`",
        "",
        "## Headers vs canônico (extrato CB MOVE)",
        "",
        f"- Status: **{header_report['header_parity']}**",
        f"- Drive: `{header_report['headers_drive']}`",
        f"- Faltando: `{header_report['missing_vs_canon']}`",
        f"- Extra: `{header_report['extra_vs_canon']}`",
        "",
        "## Contagem de células",
        "",
        f"| Status | Qtd |",
        f"| ------ | --- |",
        f"| ok | {summary['contagem']['ok']} |",
        f"| diverge | {summary['contagem']['diverge']} |",
        f"| ausente no sistema | {summary['contagem']['ausente_sistema']} |",
        f"| ausente no Drive | {summary['contagem']['ausente_drive']} |",
        "",
        f"Linhas Drive: **{summary['drive_linhas']}** · Sistema: **{summary['sistema_linhas']}** · API: {'sim' if summary['sistema_disponivel'] else 'não (sem service role)'}",
        "",
        "## Divergências (não-ok)",
        "",
        "| Paciente | Campo | Drive | Sistema | Status |",
        "| -------- | ----- | ----- | ------- | ------ |",
    ]
    for d in diffs:
        if d.status == "ok":
            continue
        lines.append(
            f"| {d.paciente} | {d.campo} | {d.drive!s} | {d.sistema!s} | {d.status} |"
        )

    lines += ["", "## Auditoria de headers — todas as abas", ""]
    for name, rep in all_headers.items():
        lines.append(
            f"- **{name}**: {rep['header_parity']} · missing={rep['missing_vs_canon']} · extra={rep['extra_vs_canon']}"
        )

    out_md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    latest = OUT / "LATEST.md"
    latest.write_text(out_md.read_text(encoding="utf-8"), encoding="utf-8")
    (OUT / "headers-todas-abas.json").write_text(
        json.dumps(all_headers, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(json.dumps({"out_md": str(out_md), "out_json": str(out_json), "contagem": summary["contagem"], "header": header_report}, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
