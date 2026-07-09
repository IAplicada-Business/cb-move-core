#!/usr/bin/env python3
"""
Backfill frequencia_atendimento e dias_semana a partir da planilha Relatório Financeiro 2026.

Uso:
  python scripts/backfill-frequencia-planilha.py
  python scripts/backfill-frequencia-planilha.py --file caminho.xlsx
  python scripts/backfill-frequencia-planilha.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import requests

from load_app_env import load_app_env

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = Path(__file__).resolve().parent / "drive_import"
SHEET_ID = "1gaAM58gv9zWCdyF8-Kp7di6WvpC9EFgeS5FsSpxV8Y0"
XLSX_NAME = "relatorio_financeiro_2026.xlsx"

ABA_MES: dict[str, int] = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARÇO": 3,
    "MARCO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}


@dataclass
class LinhaPlanilha:
    paciente_nome: str
    mes: int
    ano: int
    frequencia: str | None
    dias_semana: str | None
    valor: float | None
    plano: str
    qtd_sessoes: int | None
    situacao: str


def norm_nome(n: str) -> str:
    return (
        unicodedata.normalize("NFD", n.strip().lower())
        .encode("ascii", "ignore")
        .decode("ascii")
    )


def parse_valor(v) -> float | None:
    if isinstance(v, (int, float)) and v > 0:
        return float(v)
    s = str(v or "").replace("R$", "").replace(" ", "").replace(",", ".")
    try:
        f = float(s)
        return f if f > 0 else None
    except ValueError:
        return None


def deve_ignorar(nome: str, plano: str, situacao: str) -> bool:
    if not nome or len(nome) < 3:
        return True
    if nome == "Nome do Paciente":
        return True
    if plano == "*****":
        return True
    if "sem cobran" in situacao.lower():
        return True
    return False


def infer_servico_tag(frequencia: str) -> str | None:
    f = frequencia.lower()
    if "triplo" in f:
        return "triplo"
    if "duplo" in f:
        return "duplo"
    if "simples" in f:
        return "simples"
    return None


def download_planilha(dest: Path) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    session = requests.Session()
    r = session.get(url, timeout=120, allow_redirects=True)
    if "text/html" in r.headers.get("content-type", "") and "confirm=" in r.text:
        m = re.search(r"confirm=([0-9A-Za-z_]+)", r.text)
        if m:
            r = session.get(
                f"https://drive.google.com/uc?export=download&confirm={m.group(1)}&id={SHEET_ID}",
                timeout=120,
            )
    r.raise_for_status()
    dest.write_bytes(r.content)
    return dest


def ler_planilha(path: Path, ano: int = 2026) -> list[LinhaPlanilha]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    linhas: list[LinhaPlanilha] = []

    for sheet_name in wb.sheetnames:
        mes = ABA_MES.get(sheet_name.strip().upper())
        if not mes:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        for row in rows[2:]:
            cells = list(row) + [None] * (10 - len(row))
            nome = str(cells[0] or "").strip()
            frequencia = str(cells[2] or "").strip() or None
            dias = str(cells[3] or "").strip() or None
            qtd_raw = cells[4]
            plano = str(cells[5] or "").strip()
            valor = parse_valor(cells[7])
            situacao = str(cells[9] or "").strip()
            if deve_ignorar(nome, plano, situacao):
                continue
            qtd = None
            if qtd_raw is not None and str(qtd_raw).strip():
                try:
                    qtd = int(float(str(qtd_raw).replace(",", ".")))
                except ValueError:
                    qtd = None
            if not frequencia and not dias:
                continue
            linhas.append(
                LinhaPlanilha(
                    paciente_nome=nome,
                    mes=mes,
                    ano=ano,
                    frequencia=frequencia,
                    dias_semana=dias,
                    valor=valor,
                    plano=plano,
                    qtd_sessoes=qtd,
                    situacao=situacao,
                )
            )
    wb.close()
    return linhas


def _headers(service_key: str) -> dict[str, str]:
    return {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
    }


def fetch_all_cobrancas(base: str, service_key: str) -> list[dict]:
    rows: list[dict] = []
    offset = 0
    page = 1000
    while True:
        url = (
            f"{base}/rest/v1/cobrancas"
            "?select=id,paciente_id,valor,competencia_mes,competencia_ano,servico,regime,"
            "frequencia_atendimento,dias_semana,qtd_sessoes,"
            "pacientes(nome)"
            f"&competencia_ano=eq.2026"
            f"&order=created_at.asc"
            f"&offset={offset}&limit={page}"
        )
        req = urllib.request.Request(url, headers=_headers(service_key), method="GET")
        with urllib.request.urlopen(req, timeout=60) as resp:
            batch = json.loads(resp.read().decode())
        if not batch:
            break
        rows.extend(batch)
        if len(batch) < page:
            break
        offset += page
    return rows


def fetch_pacientes(base: str, service_key: str) -> list[dict]:
    url = f"{base}/rest/v1/pacientes?select=id,nome,frequencia_atendimento,dias_semana"
    req = urllib.request.Request(url, headers=_headers(service_key), method="GET")
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode())


def patch_row(base: str, service_key: str, table: str, row_id: str, payload: dict, dry_run: bool) -> bool:
    if dry_run:
        return True
    url = f"{base}/rest/v1/{table}?id=eq.{row_id}"
    data = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=data, headers=_headers(service_key), method="PATCH")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return 200 <= resp.status < 300
    except urllib.error.HTTPError as e:
        print(f"  ERRO PATCH {table} {row_id}: {e.code} {e.read().decode()[:200]}", file=sys.stderr)
        return False


def infer_regime(plano: str) -> str:
    p = plano.strip().lower()
    return "por_sessao" if "sess" in p else "mensalista"


def match_cobranca(linha: LinhaPlanilha, candidatos: list[dict]) -> dict | None:
    if not candidatos:
        return None
    regime = infer_regime(linha.plano)
    candidatos = [c for c in candidatos if (c.get("regime") or "mensalista") == regime] or candidatos

    tag = infer_servico_tag(linha.frequencia or "")
    if linha.valor is not None:
        por_valor = [
            c
            for c in candidatos
            if abs(float(c["valor"]) - linha.valor) < 0.02
        ]
        if len(por_valor) == 1:
            return por_valor[0]
        candidatos = por_valor or candidatos
    if tag:
        por_tag = [
            c
            for c in candidatos
            if tag in str(c.get("servico") or "").lower()
        ]
        if len(por_tag) == 1:
            return por_tag[0]
        if por_tag:
            candidatos = por_tag
    if len(candidatos) == 1:
        return candidatos[0]
    if linha.valor is None and tag:
        por_tag = [c for c in candidatos if tag in str(c.get("servico") or "").lower()]
        if len(por_tag) == 1:
            return por_tag[0]
    return None


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="Caminho local do xlsx (senão baixa do Google)")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    load_app_env()
    import os

    base = os.environ.get("SUPABASE_URL") or os.environ.get("VITE_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not base or not key:
        print("Defina SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY em .env.app", file=sys.stderr)
        return 1

    xlsx_path = Path(args.file) if args.file else OUT_DIR / XLSX_NAME
    if not xlsx_path.is_file():
        print(f"Baixando planilha → {xlsx_path}")
        download_planilha(xlsx_path)
    else:
        print(f"Usando planilha local: {xlsx_path}")

    linhas = ler_planilha(xlsx_path)
    print(f"Linhas com frequência/dias na planilha: {len(linhas)}")

    cobrancas = fetch_all_cobrancas(base, key)
    pacientes = fetch_pacientes(base, key)
    paciente_by_norm = {norm_nome(p["nome"]): p for p in pacientes}

    cobrancas_por_chave: dict[tuple[str, int, int], list[dict]] = {}
    for c in cobrancas:
        nome = c.get("pacientes", {}) or {}
        nome_str = nome.get("nome") if isinstance(nome, dict) else None
        if not nome_str:
            continue
        chave = (norm_nome(nome_str), c["competencia_mes"], c["competencia_ano"])
        cobrancas_por_chave.setdefault(chave, []).append(c)

    atualizadas_cob = 0
    sem_match = 0
    ja_ok = 0

    for linha in linhas:
        chave = (norm_nome(linha.paciente_nome), linha.mes, linha.ano)
        candidatos = cobrancas_por_chave.get(chave, [])
        cob = match_cobranca(linha, candidatos)
        if not cob:
            sem_match += 1
            continue

        payload: dict = {}
        if linha.frequencia and (cob.get("frequencia_atendimento") or "").strip() != linha.frequencia:
            payload["frequencia_atendimento"] = linha.frequencia
        if linha.dias_semana and (cob.get("dias_semana") or "").strip() != linha.dias_semana:
            payload["dias_semana"] = linha.dias_semana
        if linha.qtd_sessoes and not cob.get("qtd_sessoes"):
            payload["qtd_sessoes"] = linha.qtd_sessoes

        if not payload:
            ja_ok += 1
            continue

        if patch_row(base, key, "cobrancas", cob["id"], payload, args.dry_run):
            atualizadas_cob += 1

    # Pacientes: usa o registro mais recente (maior mês) por nome
    latest_paciente: dict[str, LinhaPlanilha] = {}
    for linha in linhas:
        n = norm_nome(linha.paciente_nome)
        prev = latest_paciente.get(n)
        if not prev or linha.mes > prev.mes:
            latest_paciente[n] = linha

    atualizados_pac = 0
    for norm, linha in latest_paciente.items():
        pac = paciente_by_norm.get(norm)
        if not pac:
            continue
        payload: dict = {}
        if linha.frequencia and (pac.get("frequencia_atendimento") or "").strip() != linha.frequencia:
            payload["frequencia_atendimento"] = linha.frequencia
        if linha.dias_semana and (pac.get("dias_semana") or "").strip() != linha.dias_semana:
            payload["dias_semana"] = linha.dias_semana
        if not payload:
            continue
        if patch_row(base, key, "pacientes", pac["id"], payload, args.dry_run):
            atualizados_pac += 1

    modo = "DRY-RUN" if args.dry_run else "APPLY"
    print(f"\n=== Backfill {modo} ===")
    print(f"Cobranças atualizadas: {atualizadas_cob}")
    print(f"Cobranças já preenchidas: {ja_ok}")
    print(f"Linhas sem match: {sem_match}")
    print(f"Pacientes atualizados: {atualizados_pac}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
